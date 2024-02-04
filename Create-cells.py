import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
import xml.etree.ElementTree as ET

zomboid_path = ''
output_path = ''
root = tk.Tk()
root.withdraw()
workshop_box = messagebox.askokcancel('Workshop folder','Please select your Project Zomboid workshop folder called "108600", which is usually located in "C:\Program Files (x86)\Steam\steamapps\workshop\content"')
if workshop_box:
    zomboid_path = filedialog.askdirectory(initialdir='C:\Program Files (x86)\Steam\steamapps\workshop\content',mustexist=True,title='Please select your Project Zomboid workshop folder: 108600')
    if zomboid_path == '':
        quit()
else:
    quit()
output_box = messagebox.askokcancel('Output path','Please select where you want to save your XLSX file')
if output_box:
    output_path = filedialog.askdirectory(mustexist=True,title='Please select your save location')
    if output_path == '':
        quit()
else:
    quit()
options_box = messagebox.askyesnocancel('Save Options' , 'Do you want to include tilemaps of cells which are conflicting? NOTE: Increases program runtime a bit.')
if options_box:
    extended_box = messagebox.askyesnocancel('Full Search?' , 'Do you want to include all tilesmaps? WARNING: The Script will take a lot of time to complete depending on your workshopfolder. On a 5800X3D it took me around 5 minutes to generate a 28MB file for 100 mapmods.')
    if extended_box == '':
        quit()
if options_box == '':
    quit()
# Do some math magic for linegeneration
def bresenham_line(x0, y0, x1, y1):
    points = []
    dx = abs(x1 - x0)
    dy = abs(y1 - y0)
    x, y = x0, y0
    sx = -1 if x0 > x1 else 1
    sy = -1 if y0 > y1 else 1
    if dx > dy:
        err = dx / 2.0
        while x != x1:
            points.append((x, y))
            err -= dy
            if err < 0:
                y += sy
                err += dx
            x += sx
    else:
        err = dy / 2.0
        while y != y1:
            points.append((x, y))
            err -= dx
            if err < 0:
                x += sx
                err += dy
            y += sy        
    points.append((x, y))
    return points

# create coordinates between two coordinates
def insert_intermediate_coords(coordinate):
    result = []
    for i in range(len(coordinate) - 1):
        result.extend(bresenham_line(*coordinate[i], *coordinate[i+1]))
    return result

def stop():
    quit()

def find_files_with_coordinates(root_folder):

    # Create a dictionary to store the folder names
    folders = {}
    foldersconf = {}
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Mapcells'
    
    # Walk through the root folder and its subfolders
    for foldername, subfolders, filenames in os.walk(root_folder):
        # Loop through the filenames
        for filename in filenames:
        # Check if the filename matches the pattern
            if filename.endswith('.lotheader'):
                # Extract the coordinates from the filename
                coords = filename.split('.')[0].split('_')
                # Get the folder name
                folder = os.path.basename(foldername)
                # Add the folder name to the dictionary
                if (coords[0], coords[1]) in folders:
                    foldersconf[(coords[0], coords[1])] = folder
                    folders[(coords[0], coords[1])] += ' +' + folder
                else:
                    if extended_box:
                        foldersconf[(coords[0], coords[1])] = folder
                    folders[(coords[0], coords[1])] = folder
    # Create the lables and run through each cell
    ws.cell(row=1, column=2).value = 'X'
    ws.cell(row=2, column=1).value = 'Y'
    for i in range(0, 101):
        ws.cell(row=1, column=i+3).value = i
        ws.cell(row=i+3, column=1).value = i
        for j in range(0, 101):
            cell = folders.get((str(j), str(i)), '')
            if ' +' in cell:
                cell = '!' + cell
                ws.cell(row=i+3, column=j+3).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
            ws.cell(row=i+3, column=j+3).value = cell
            ws.cell(row=i+3, column=j+3).alignment = Alignment(vertical='justify')
    if options_box:
        pages_edited = []     
        page_edited_count = {}
        folderschecked = 0
        # Create a progressbar
        rootup = tk.Tk()
        rootup.geometry('300x120')
        rootup.title('Cells to Tiles')
        progress = ttk.Progressbar(
            rootup,
            orient='horizontal',
            mode='determinate',
            length=280
        )
        progress.grid(column=0, row=0, columnspan=2, padx=10, pady=20)
        progress['value'] = 0
        progress_label = ttk.Label(rootup, text=str(folderschecked) + '/' + str(len(foldersconf)) + 'checked')
        progress_label.grid(column=0, row=1, columnspan=2)
        progress.update_idletasks()
        progress.update()
        # Walk through the root folder and its subfolders
        for foldername, subfolders, filenames in os.walk(root_folder):
            # Loop through the filenames
            for filename in filenames:
            # Check if the filename matches the pattern
                if filename == 'worldmap.xml':
                    xml_file = os.path.join(foldername, 'worldmap.xml')

                    # Parse the worldmap.xml file
                    tree = ET.parse(xml_file)
                    root = tree.getroot()
                    # Get the folder name
                    folder = os.path.basename(foldername)

                    # Loop through the cells in the worldmap.xml file
                    for cell in root.iter('cell'):
                        # Get the X and Y coordinates
                        x = max(0, int(cell.get('x')))
                        y = max(0, int(cell.get('y')))
                        map = wb['Mapcells']
                        # Check if the cell is a conflicting one and full search is not enabled
                        if '!' not in map.cell(row=y+3, column=x+3).value and not extended_box:
                            continue
                        folderschecked += 1
                        try:
                            progress['value'] = folderschecked/len(foldersconf)*100
                            progress_label['text'] = (str(folderschecked) + '/' + str(len(foldersconf)) + ' checked')
                            progress.update_idletasks()
                            progress.update()
                        except:
                            quit()
                        # Select the fitting worksheet or create one if there is none
                        try:
                            ws = wb['Cell ({}, {})'.format(x, y)]
                        except:
                            ws = wb.create_sheet(title='Cell ({}, {})'.format(x, y))
                        # Create the legend
                        if folder not in pages_edited:
                            pages_edited.append(folder)
                            page_edited_count['Cell ({}, {})'.format(x, y)] = 1
                            ws.cell(row=int(page_edited_count['Cell ({}, {})'.format(x, y)])+3, column=1).value = folder + '=' + str(page_edited_count['Cell ({}, {})'.format(x, y)])
                            ws.cell(row=int(page_edited_count['Cell ({}, {})'.format(x, y)])+3, column=1).fill = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
                            ws.cell(row=int(page_edited_count['Cell ({}, {})'.format(x, y)])+3, column=1).font = Font(color='FFFFFFFF')
                        elif 'Cell ({}, {})'.format(x, y) not in page_edited_count:
                            page_edited_count['Cell ({}, {})'.format(x, y)] = 1
                            ws.cell(row=int(page_edited_count['Cell ({}, {})'.format(x, y)])+3, column=1).value = folder + '=' + str(page_edited_count['Cell ({}, {})'.format(x, y)])
                            ws.cell(row=int(page_edited_count['Cell ({}, {})'.format(x, y)])+3, column=1).fill = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
                            ws.cell(row=int(page_edited_count['Cell ({}, {})'.format(x, y)])+3, column=1).font = Font(color='FFFFFFFF')
                        elif page_edited_count.get('Cell ({}, {})'.format(x, y)) == 1:
                            page_edited_count['Cell ({}, {})'.format(x, y)] += 1
                            ws.cell(row=int(page_edited_count['Cell ({}, {})'.format(x, y)])+3, column=1).value = folder + '=' + str(page_edited_count['Cell ({}, {})'.format(x, y)])
                            ws.cell(row=int(page_edited_count['Cell ({}, {})'.format(x, y)])+3, column=1).fill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
                            ws.cell(row=int(page_edited_count['Cell ({}, {})'.format(x, y)])+3, column=1).font = Font(color='FFFFFFFF')
                        elif page_edited_count.get('Cell ({}, {})'.format(x, y)) == 2:
                            page_edited_count['Cell ({}, {})'.format(x, y)] += 1
                            ws.cell(row=int(page_edited_count['Cell ({}, {})'.format(x, y)])+3, column=1).value = folder + '=' + str(page_edited_count['Cell ({}, {})'.format(x, y)])
                            ws.cell(row=int(page_edited_count['Cell ({}, {})'.format(x, y)])+3, column=1).fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
                            ws.cell(row=int(page_edited_count['Cell ({}, {})'.format(x, y)])+3, column=1).font = Font(color='FFFFFFFF')
                        else:
                            page_edited_count['Cell ({}, {})'.format(x, y)] += 1
                            ws.cell(row=int(page_edited_count['Cell ({}, {})'.format(x, y)])+3, column=1).value = folder + '=' + str(page_edited_count['Cell ({}, {})'.format(x, y)])
                            ws.cell(row=int(page_edited_count['Cell ({}, {})'.format(x, y)])+3, column=1).fill = PatternFill(start_color='FF00FFFF', end_color='FF00FFFF', fill_type='solid')
                            ws.cell(row=int(page_edited_count['Cell ({}, {})'.format(x, y)])+3, column=1).font = Font(color='FFFFFFFF')
                        ws.cell(row=1, column=1).value = 'Legend:'
                        ws.cell(row=2, column=1).value = 'Conflicts'
                        ws.cell(row=2, column=1).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

                        # Write the table headers
                        ws.cell(row=1, column=3).value = 'X'
                        ws.cell(row=2, column=2).value = 'Y'

                        # Write the X values to the worksheet
                        for i in range(0, 306):
                            ws.cell(row=1, column=i+4).value = i

                        # Write the Y values to the worksheet
                        for i in range(0, 306):
                            ws.cell(row=i+3, column=2).value = i

                        # Loop through the points in the cell
                        coord_iteration = 0
                        for coordinates in cell.iter('coordinates'):
                            coord_iteration += 1
                            pbuffer = []
                            for point in coordinates.iter('point'):
                                # Get the X and Y coordinates
                                xp = max(0, int(point.get('x')))
                                yp = max(0, int(point.get('y')))
                                # Write the X and Y coordinates to the worksheet
                                if ws.cell(row=yp+3, column=xp+4).value is None:
                                    ws.cell(row=yp+3, column=xp+4).value = str(page_edited_count['Cell ({}, {})'.format(x, y)])
                                    ws.cell(row=yp+3, column=xp+4).font = Font(color='FFFFFFFF')
                                else:
                                    ws.cell(row=yp+3, column=xp+4).value += '+' + str(page_edited_count['Cell ({}, {})'.format(x, y)])
                                    ws.cell(row=yp+3, column=xp+4).font = Font(color='FFFFFFFF')
                                # Buffer all points in each cordinate
                                pbuffer.append((xp,yp))
                            pbuffer.append(pbuffer[0])
                            # Use the math magic to fill in missing coordinates of the buffer then paint each coordinate in the corresponding color
                            full_coords = insert_intermediate_coords(pbuffer)
                            for coordinate in full_coords:
                                if page_edited_count.get('Cell ({}, {})'.format(x, y)) == 1:
                                    ws.cell(row=coordinate[1]+3, column=coordinate[0]+4).fill = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
                                elif page_edited_count.get('Cell ({}, {})'.format(x, y)) == 2:
                                    if ws.cell(row=coordinate[1]+3, column=coordinate[0]+4).fill != PatternFill():
                                        ws.cell(row=coordinate[1]+3, column=coordinate[0]+4).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                                    else:
                                        ws.cell(row=coordinate[1]+3, column=coordinate[0]+4).fill = PatternFill(start_color='FF0000FF', end_color='FF0000FF', fill_type='solid')
                                elif page_edited_count.get('Cell ({}, {})'.format(x, y)) == 3:
                                    if ws.cell(row=coordinate[1]+3, column=coordinate[0]+4).fill != PatternFill():
                                        ws.cell(row=coordinate[1]+3, column=coordinate[0]+4).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                                    ws.cell(row=coordinate[1]+3, column=coordinate[0]+4).fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
                                elif ws.cell(row=coordinate[1]+3, column=coordinate[0]+4).fill != PatternFill():
                                        ws.cell(row=coordinate[1]+3, column=coordinate[0]+4).fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                                else:
                                    ws.cell(row=coordinate[1]+3, column=coordinate[0]+4).fill = PatternFill(start_color='FF00FFFF', end_color='FF00FFFF', fill_type='solid')

        progress['value'] = 100
        progress_label['text'] = ('Saving file')
        progress.update_idletasks()
        progress.update()
        for sheet in wb.sheetnames:
            if sheet != 'Mapcells':
                ws = wb[sheet]
                # Remove all sheets that don't have a second mod entry
                if ws.cell(row=5, column=1).value is None and not extended_box:
                    wb.remove(ws)
                # Change the cell width in all sheets except in the main one
                for col in range(2, ws.max_column + 1):
                    ws.column_dimensions[ws.cell(1, col).column_letter].width = 4
        # Put the main sheet at the front again
        wb._sheets.sort(key=lambda ws:ws.title)
        wb.move_sheet('Mapcells', -(len(wb.sheetnames)-1))
        # Save the workbook
    wb.save(output_path + '/Mapcells.xlsx')

# Call the function with the zomboid folder
find_files_with_coordinates(zomboid_path)
exit_message = messagebox.showinfo('Done','You can find your XLSX File here: ' + output_path + '/Mapcells.xlsx')
